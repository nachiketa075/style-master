"""
Builds the fixed-header Excel workbook (Rule 9, Rule 18) and handles
opening + maximizing it in Excel after conversion (Rule 21).
"""

from __future__ import annotations

import datetime
import os
import time

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import config

_HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_THIN = Side(style="thin", color="B7B7B7")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP_COLUMNS = {"Component Details", "EXTRA COMPONENTS", "Metal Details"}

# These columns are free-form measurement STRINGS, never numbers -
# e.g. MM Size can be "1.8x1.8" or "0.90", and weight columns like
# "All Total Wt" can be "0.1650" where the trailing zero is significant
# (a value python's own str() type already preserves exactly, since
# nothing in this pipeline ever runs int()/float() on OCR output).
# Forcing the cell's Excel number format to Text ("@") is
# belt-and-suspenders: it stops Excel itself from ever
# reinterpreting/reformatting the stored value on open (e.g. silently
# dropping "0.1650" down to "0.165").
_TEXT_FORMAT_COLUMNS = {"MM Size", "Sieve Size", "Per Stone Wt", "Total Wt", "All Total Wt"}


def build_workbook(all_rows: list[dict]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = config.WORKSHEET_NAME
    _write_sheet(ws, config.FIXED_HEADERS, all_rows)

    if config.DEBUG_MODE:
        # Debug sheet content is populated by main.py when available;
        # left as an empty, clearly-labelled sheet otherwise.
        wb.create_sheet(config.DEBUG_WORKSHEET_NAME)

    return wb


def _write_sheet(ws: Worksheet, headers: list[str], all_rows: list[dict]) -> None:
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER

    for row_dict in all_rows:
        ws.append([row_dict.get(h, "") for h in headers])

    max_row = ws.max_row
    max_col = len(headers)
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = _BORDER
            if cell.row > 1:
                header_name = headers[cell.column - 1]
                wrap = header_name in _WRAP_COLUMNS
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=wrap
                )
                if header_name in _TEXT_FORMAT_COLUMNS:
                    cell.number_format = "@"

    ws.freeze_panes = "A2"
    if max_row >= 1 and max_col >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    for col_idx, header in enumerate(headers, start=1):
        length = len(str(header))
        for row_idx in range(2, max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                length = max(length, len(str(val)))
        width = min(max(length + 2, 10), 45)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_debug_sheet(wb: Workbook, debug_rows: list[dict]) -> None:
    if config.DEBUG_WORKSHEET_NAME not in wb.sheetnames:
        wb.create_sheet(config.DEBUG_WORKSHEET_NAME)
    ws = wb[config.DEBUG_WORKSHEET_NAME]
    headers = ["Image", "Text", "Left", "Top", "Width", "Height", "Confidence"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in debug_rows:
        ws.append([r.get(h, "") for h in headers])
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16


def unique_output_path(output_dir: str) -> str:
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    base_name = f"{config.OUTPUT_FILENAME_PREFIX}_{timestamp}.xlsx"
    path = os.path.join(output_dir, base_name)
    counter = 1
    while os.path.exists(path):
        path = os.path.join(
            output_dir, f"{config.OUTPUT_FILENAME_PREFIX}_{timestamp}_{counter}.xlsx"
        )
        counter += 1
    return path


def save_workbook(wb: Workbook, path: str) -> str:
    wb.save(path)
    if not os.path.exists(path):
        raise IOError("Excel file was not created on disk.")
    return path


def open_and_maximize_excel(path: str, timeout: float = 15.0) -> bool:
    """Launch the exported workbook in Excel, force its OS window to a
    fully maximized state, and confirm it is actually visible before
    returning.

    Returns True once this specific workbook's Excel window is
    confirmed open, visible, and maximized; False if that could not
    be confirmed within `timeout` seconds (the workbook itself is
    already saved to disk either way - this function never touches
    the file and never fails the export).

    Uses the real Win32 API (ShowWindow/SW_MAXIMIZE on the window
    handle) rather than only the COM WindowState property, which was
    found not to reliably resize the actual OS window in every
    environment. Windows are matched by title against this workbook's
    filename (via EnumWindows), not just "the" Excel.Application COM
    singleton, so a workbook the user already had open in another
    Excel window is left alone and never touched.

    Never calls Excel.Quit() or closes any Excel window - on success
    this intentionally leaves Excel running, whether it was a
    pre-existing instance or one just launched for this file.
    """
    os.startfile(path)  # noqa: S606 - intentional, user-selected local file

    try:
        import win32con
        import win32gui
    except ImportError:
        return False

    target_stem = os.path.splitext(os.path.basename(path))[0].lower()

    def _find_window() -> int | None:
        found: list[int] = []

        def _enum_handler(hwnd, _):
            if found:
                return
            if not win32gui.IsWindowVisible(hwnd):
                return
            try:
                if win32gui.GetClassName(hwnd) != "XLMAIN":
                    return
            except Exception:
                return
            if target_stem in win32gui.GetWindowText(hwnd).lower():
                found.append(hwnd)

        win32gui.EnumWindows(_enum_handler, None)
        return found[0] if found else None

    deadline = time.time() + timeout
    hwnd = None
    while time.time() < deadline and hwnd is None:
        hwnd = _find_window()
        if hwnd is None:
            time.sleep(0.3)

    if hwnd is None:
        return False

    confirm_deadline = time.time() + timeout
    while time.time() < confirm_deadline:
        try:
            win32gui.ShowWindow(hwnd, win32con.SW_MAXIMIZE)
            win32gui.SetForegroundWindow(hwnd)
        except Exception:
            pass
        if win32gui.IsWindowVisible(hwnd) and win32gui.GetWindowPlacement(hwnd)[1] == win32con.SW_SHOWMAXIMIZED:
            return True
        time.sleep(0.3)

    return False
